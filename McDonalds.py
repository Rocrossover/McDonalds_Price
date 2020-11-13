##import requests
##from urllib.request import Request, urlopen
##import  xlsxwriter
from bs4 import BeautifulSoup as soup
from selenium import webdriver
import re
import datetime
import openpyxl

##request = Request(url, headers={'User-Agent': 'Mozilla/5.0'})
##web = urlopen(request).read()
##web_soup = soup(web,'html.parser')

url = 'https://www.skipthedishes.com/mcdonalds-dundurn-street-south'
driver = webdriver.Chrome()
driver.get(url)
html = driver.execute_script("return document.documentElement.outerHTML")
sel_soup = soup(html,'html.parser')
menu_info = sel_soup.findAll("div",{"class":"styles__TextWrapper-sc-1xl58bi-9 cxewEF"})

name = []
price = []

x= datetime.datetime.now()
X = x.strftime("%b")
Y = x.strftime("%d")
Z = X+" "+Y

##open a new workbook
##McD = openpyxl.Workbook()
McD = openpyxl.load_workbook('McD_Price.xlsx')
WS = McD.create_sheet(Z)
NAME = WS.cell(row=1, column=1, value="Name")
PRICE = WS.cell(row=1, column=2, value="Price")

##Menu Price
for i in range(len(menu_info)):
    info = menu_info[i].find("span",{"itemprop":"name"})
    if info == None :
        name.append('')
        price.append('')
        pass
    else:
        name.append(info.string)
        price.append(info.parent.parent.next_sibling.next_sibling.find("h3").string)
        try:
            WS.cell(row=i+2,column=1).value = name[i]
            WS.cell(row=i+2,column=2).value = price[i]
            print(i,name[i],price[i])
        except:
            pass

##Remove empty rows in excel
index_row = []
for i in range(1, WS.max_row):
    if WS.cell(i, 1).value is None:
        # collect indexes of rows
        index_row.append(i)

for row_del in range(len(index_row)):
    WS.delete_rows(idx=index_row[row_del], amount=1)
    # exclude offset of rows through each iteration
    index_row = list(map(lambda k: k - 1, index_row))


McD.save('McD_Price.xlsx')
McD.close()
driver.close()

print("Successful!")
